import csv
import io
from datetime import datetime
from typing import Any, Dict, List, Tuple

from openpyxl import load_workbook

from parser import ImportOptions, parse_for_staging
from rule_manager import infer_category, is_known_category, is_known_l1, l2_to_l1


REQUIRED_COLUMNS = ["日期", "金额", "类型明细", "交易方向"]
OPTIONAL_COLUMNS = ["日记", "类型", "一级类型", "旅游标识", "旅游标签"]


def _to_float(val: Any):
    if val is None:
        return None
    text = str(val).strip().replace(",", "")
    if text == "":
        return None
    try:
        return float(text)
    except Exception:
        return None


def _normalize_date(val: Any):
    if val is None:
        return None
    text = str(val).strip().replace("/", "-").replace(".", "-")
    if text == "":
        return None
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(text, fmt).strftime("%Y-%m-%d")
        except Exception:
            pass
    parts = text.split("-")
    if len(parts) == 3 and all(p.isdigit() for p in parts):
        y, m, d = int(parts[0]), int(parts[1]), int(parts[2])
        if 2000 <= y <= 2100 and 1 <= m <= 12 and 1 <= d <= 31:
            return "{:04d}-{:02d}-{:02d}".format(y, m, d)
    return None


def _normalize_direction(val: Any):
    text = str(val).strip()
    if text in ("收入", "in", "IN", "income", "Income"):
        return "收入"
    if text in ("支出", "out", "OUT", "expense", "Expense"):
        return "支出"
    return None


def _normalize_travel_flag(val: Any):
    text = str(val).strip().lower()
    return text in ("1", "true", "yes", "y", "是", "旅游")


def _build_stage_row(row_index: int, row: Dict[str, Any]) -> Dict[str, Any]:
    bill_date = _normalize_date(row.get("日期"))
    amount = _to_float(row.get("金额"))
    direction = _normalize_direction(row.get("交易方向"))
    detail = str(row.get("类型明细", "")).strip()
    note = str(row.get("日记", "")).strip()
    category_l2 = str(row.get("类型", "")).strip()
    category_l1_raw = str(row.get("一级类型", "")).strip()
    is_travel = _normalize_travel_flag(row.get("旅游标识"))
    travel_tag = str(row.get("旅游标签", "")).strip()

    errors = []
    if bill_date is None:
        errors.append("日期格式错误")
    if amount is None:
        errors.append("金额格式错误")
        amount = 0.0
    if direction is None:
        errors.append("交易方向必须是收入或支出")
        direction = "支出"
    if not detail:
        errors.append("类型明细不能为空")

    explicit_raw = ""
    category_unknown = False
    is_inc = direction == "收入"

    if not category_l2:
        cat_l1, cat_l2 = infer_category(detail, is_inc)
        category_l1_raw = cat_l1
        category_l2 = cat_l2
    else:
        # 兼容旧习惯：用户只填“类型”时，默认按一级类型处理
        # 例如“生活支出”会被视为 category_l1=生活支出, category_l2=生活支出
        if is_known_l1(category_l2, is_inc):
            explicit_raw = ""
            category_unknown = False
            category_l1_raw = category_l2
        elif is_known_category(category_l2, is_inc):
            explicit_raw = ""
            category_unknown = False
            if not category_l1_raw:
                category_l1_raw = l2_to_l1(category_l2, is_inc) or category_l2
        else:
            category_unknown = True
            explicit_raw = category_l2

    return {
        "row_index": row_index,
        "bill_date": bill_date,
        "amount": amount,
        "detail": detail,
        "note": note,
        "direction": direction,
        "category_l1": category_l1_raw,
        "category": category_l2,
        "explicit_category_raw": explicit_raw if category_unknown else "",
        "category_unknown": category_unknown,
        "is_travel": is_travel,
        "travel_tag": travel_tag,
        "is_valid": len(errors) == 0,
        "error_msg": "；".join(errors),
    }


def _validate_columns(headers: List[str]) -> Tuple[bool, str]:
    missing = [col for col in REQUIRED_COLUMNS if col not in headers]
    if missing:
        return False, "缺少必填字段: {}".format(",".join(missing))
    return True, ""


def parse_csv_bytes(data: bytes) -> Tuple[List[Dict[str, Any]], str]:
    text = data.decode("utf-8-sig", errors="ignore")
    f = io.StringIO(text)
    reader = csv.DictReader(f)
    if not reader.fieldnames:
        return [], "CSV没有表头"
    ok, err = _validate_columns([h.strip() for h in reader.fieldnames])
    if not ok:
        return [], err

    rows = []
    for idx, row in enumerate(reader, start=1):
        rows.append(_build_stage_row(idx, row))
    return rows, ""


def parse_xlsx_bytes(data: bytes) -> Tuple[List[Dict[str, Any]], str]:
    wb = load_workbook(io.BytesIO(data), data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return [], "XLSX为空"

    headers = [str(v).strip() if v is not None else "" for v in all_rows[0]]
    ok, err = _validate_columns(headers)
    if not ok:
        return [], err

    rows = []
    for idx, raw in enumerate(all_rows[1:], start=1):
        item = {}
        for i, h in enumerate(headers):
            item[h] = raw[i] if i < len(raw) else None
        rows.append(_build_stage_row(idx, item))
    return rows, ""


def parse_input_to_staging(
    file_ext: str,
    file_bytes: bytes,
    raw_text: str,
    options: ImportOptions,
) -> Tuple[List[Dict[str, Any]], str]:
    ext = (file_ext or "").lower()
    if ext == ".csv":
        return parse_csv_bytes(file_bytes)
    if ext == ".xlsx":
        return parse_xlsx_bytes(file_bytes)
    if ext == ".txt":
        return parse_for_staging(raw_text, options), ""
    if raw_text.strip():
        return parse_for_staging(raw_text, options), ""
    return [], "暂不支持的格式，请使用 txt/csv/xlsx"
