import argparse
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook


SKIP_ROW_LABELS = {
    "",
    "日期",
    "共计",
    "本月现金流来源：",
    "日期：",
    "金额：",
    "本月结余：",
}


def _clean_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _parse_date_cell(value: object) -> Optional[date]:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _clean_text(value).replace(".", "-").replace("/", "-")
    if not text:
        return None
    parts = [p for p in text.split("-") if p]
    if len(parts) == 2 and all(p.isdigit() for p in parts):
        month, day = int(parts[0]), int(parts[1])
        return date(2000, month, day)
    if len(parts) == 3 and all(p.isdigit() for p in parts):
        year, month, day = int(parts[0]), int(parts[1]), int(parts[2])
        return date(year, month, day)
    return None


def _header_row_and_date_columns(ws) -> Tuple[int, List[Tuple[int, date]]]:
    header_row = 0
    for row_idx in range(1, min(ws.max_row, 8) + 1):
        if _clean_text(ws.cell(row=row_idx, column=2).value) == "日期":
            header_row = row_idx
            break
    if header_row == 0:
        raise ValueError(f"{ws.title} 未找到日期表头行")

    columns: List[Tuple[int, date]] = []
    for col in range(3, ws.max_column + 1, 2):
        day = _parse_date_cell(ws.cell(row=header_row, column=col).value)
        if day:
            columns.append((col, day))
    if not columns:
        raise ValueError(f"{ws.title} 未解析出任何日期列")
    return header_row, columns


def _number_value(value: object) -> Optional[float]:
    if isinstance(value, (int, float)):
        return float(value)
    text = _clean_text(value).replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _sheet_records(ws) -> Dict[str, List[str]]:
    header_row, date_cols = _header_row_and_date_columns(ws)
    by_date: Dict[str, List[str]] = defaultdict(list)
    direction = "支出"

    for row in range(header_row + 1, ws.max_row + 1):
        marker = _clean_text(ws.cell(row=row, column=1).value)
        row_label = _clean_text(ws.cell(row=row, column=2).value)
        if marker in {"收入", "流入"} or row_label in {"收入", "流入"}:
            direction = "收入"
            continue
        if row_label in SKIP_ROW_LABELS:
            continue

        for dcol, raw_day in date_cols:
            amount = _number_value(ws.cell(row=row, column=dcol + 1).value)
            if amount is None or abs(amount) < 1e-12:
                continue
            detail = _clean_text(ws.cell(row=row, column=dcol).value) or row_label
            old_l1 = row_label
            mmdd = f"{raw_day.month:02d}-{raw_day.day:02d}"
            if direction == "收入":
                line = f"{detail}；{amount:.2f}；收入；{old_l1}"
            else:
                line = f"{detail}；{amount:.2f}；{old_l1}"
            by_date[mmdd].append(line)
    return by_date


def convert_workbook_to_txt(xlsx_path: Path, txt_path: Path, note: str = "旧账导入") -> int:
    wb = load_workbook(xlsx_path, data_only=True)
    all_rows: Dict[str, List[str]] = defaultdict(list)

    for sheet_name in wb.sheetnames:
        title = _clean_text(sheet_name)
        if "年度总结" in title:
            continue
        rows = _sheet_records(wb[sheet_name])
        for mmdd, lines in rows.items():
            all_rows[mmdd].extend(lines)

    ordered_days = sorted(all_rows.keys(), key=lambda x: (int(x.split("-")[0]), int(x.split("-")[1])))
    blocks: List[str] = []
    for mmdd in ordered_days:
        blocks.append(f"{int(mmdd[:2])}-{int(mmdd[3:])}；{note}")
        blocks.extend(all_rows[mmdd])
        blocks.append("")

    txt_path.parent.mkdir(parents=True, exist_ok=True)
    txt_path.write_text("\n".join(blocks).strip() + "\n", encoding="utf-8")
    return sum(len(v) for v in all_rows.values())


def _detect_source_files(input_dir: Path, years: List[str]) -> List[Path]:
    files: List[Path] = []
    for fp in sorted(input_dir.glob("*.xlsx")):
        if any(y in fp.name for y in years):
            files.append(fp)
    return files


def main() -> None:
    parser = argparse.ArgumentParser(description="将旧版 Excel 账单批量提取为项目可导入 TXT")
    parser.add_argument("--input-dir", default="账单备份", help="旧账 Excel 所在目录（默认：账单备份）")
    parser.add_argument("--output-dir", default="账单备份/legacy_txt", help="输出 TXT 目录（默认：账单备份/legacy_txt）")
    parser.add_argument("--years", default="2020,2021,2022", help="需要处理的年份（逗号分隔）")
    args = parser.parse_args()

    input_dir = Path(args.input_dir)
    output_dir = Path(args.output_dir)
    years = [y.strip() for y in args.years.split(",") if y.strip()]

    files = _detect_source_files(input_dir, years)
    if not files:
        raise SystemExit(f"未在 {input_dir} 找到年份 {years} 对应的 xlsx 文件")

    for fp in files:
        year = "".join(ch for ch in fp.stem if ch.isdigit())[:4]
        note = f"旧账导入-{year}" if year else "旧账导入"
        out_file = output_dir / f"{fp.stem}-抽取.txt"
        count = convert_workbook_to_txt(fp, out_file, note=note)
        print(f"[OK] {fp.name} -> {out_file} ({count} 条)")


if __name__ == "__main__":
    main()
